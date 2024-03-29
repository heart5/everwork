#! python3
# encoding:utf-8
# ---
# jupyter:
#   jupytext:
#     cell_metadata_filter: -all
#     formats: ipynb,py:percent
#     notebook_metadata_filter: jupytext,-kernelspec,-jupytext.text_representation.jupytext_version
#     text_representation:
#       extension: .py
#       format_name: percent
#       format_version: '1.3'
# ---

# %%
# 学习excel表格的操作
# 0大数据分析（真元）.xlsx
# 39ed537d-73fa-4ad8-b4fd-bc6f746fb302 真元日配送图
import openpyxl, matplotlib.pyplot as plt, pandas as pd, numpy, datetime, sqlite3 as lite
from pandas.io import sql
from pylab import *

# %%
# plot中显示中文
mpl.rcParams['font.sans-serif'] = ['SimHei']
mpl.rcParams['axes.unicode_minus'] = False

# %%
def descdb(df):
    print(df.head(5))
    print(df.tail(5))
    print(df.dtypes)
    print(len(df))
    print(df.describe())

# %%
#
# 整理源数据，返回float数据列
#
def clean2f(dfc):
    intfloattypelist = [float, int, float32, int32, float64, int64]
    jiqian = []
    num = 0
    for ii in dfc:
        if (type(ii) in intfloattypelist):
            jiqian.append(ii)
            num += 1
            continue
        elif type(ii) == str:
            if len(ii) > 0:
                # print(str(num)+ii+'***',"\t")
                if ii.find("=") == 0:
                    # print(ii[1:],'\t')
                    ii = eval(ii[1:])
            else:
                print('问题：' + str(num) + '\t***' + ii + '***')
                ii = 0
        try:
            jiqian.append(float(ii))
        except Exception:
            print(str(num)+'错误：'+'***'+ii+'***')
            jiqian.append(0.0)
        num += 1
    return jiqian

# %%
def clean2d(dfc):
    dts = []
    num = 0
    for ii in dfc:
        if type(ii) == pd.DataFrame:
            dts.append(ii)
        else:
            try:
                dts.append(pd.to_datetime(ii))
            except:
                print(str(num)+'错误！'+ii+'不是一个合格的日期格式')
                exit()
        num += 1

    return dts

# %%
def collectdata():
    print(u'打开excel表格中……')
    wb = openpyxl.load_workbook(u'2017年全单统计管理.xlsm', read_only=True )
    # wb = openpyxl.load_workbook(u'2017年全单统计管理.xlsm')
    # wb = openpyxl.load_workbook(u'0大数据分析（真元）.xlsx')
    sheetlist = wb.get_sheet_names()
    print(sheetlist)
    # sheet = wb.get_active_sheet()
    # print(sheet)
    # sheet = wb.get_sheet_by_name('全单统计管理 (bak)')
    sheet = wb.get_sheet_by_name('全单统计管理')


    colnames = []
    count = 0
    bigdata = []
    for row in sheet.rows:
        # print("行宽：", len(row))
        if count == 0:
            for cell in row:
                colnames.append(cell.value)
            count += 1
            continue
        itemrow = []
        for cell in row:
            if cell.value is None:
                itemrow.append(0)
            else:
                itemrow.append(cell.value)
        bigdata.append(itemrow)
        count += 1

    # print(colnames)
    print(sheet.title)
    # print(sheet['A2'])
    # print(sheet['A3'].value)
    # print(bigdata[:20])
    df = pd.DataFrame(bigdata)
    # descdb(df)
    # df.fillna(value=0)
    # print(df[0:15])
    # print(df.columns)
    colnames = ['dingdanriqi', 'danhao', 'peihuoren', 'peihuozhunque', 'yewuzhuguan', 'zhongduanbianma',
                'zhongduanmingcheng', 'jiqian', 'songhuojine', 'shishoujine', 'shoukuanfangshi', 'tuihuojine',
                'kehujushou', 'wuhuojine', 'shaopeijine', 'peicuoweiyao', 'songdariqi', 'cheliang',
                'songhuoren', 'jinehedui', 'shoukuanriqi', 'shoukuanren', 'jushoupinxiang', 'shaopei']
    df.columns=colnames
    cnx = lite.connect('quandan.db')

    # df = collectdata()
    # sql_df=df.loc[:,['dingdanriqi','peihuozhunque', 'yewuzhuguan', 'zhongduanbianma', 'songhuojine',
    #                  'shishoujine', 'songdariqi', 'songhuoren', 'shoukuanriqi', 'songdatianshu',
    #                  'shoukuantianshu']]
    # df.to_sql(name='sources', con=cnx, schema=sql_df, if_exists='replace', chunksize=1000)
    df = pd.read_sql_query('select * from sources', cnx)
    descdb(df)
    df['peihuozhunque'] = df['peihuozhunque'].astype(int)

    df['songdariqi'] = clean2d(df['songdariqi'])
    df['shoukuanriqi'] = clean2d(df['shoukuanriqi'])

    df['jiqian'] = clean2f(df['jiqian'])
    df['jiqian'] = df['jiqian'].astype(float)
    df['songhuojine'] = clean2f(df['songhuojine'])
    df['songhuojine'] = df['songhuojine'].astype(float)
    df['shishoujine'] = clean2f(df['shishoujine'])
    df['shishoujine'] = df['shishoujine'].astype(float)
    df['tuihuojine'] = clean2f(df['tuihuojine'])
    df['tuihuojine'] = df['tuihuojine'].astype(float)
    df['kehujushou'] = clean2f(df['kehujushou'])
    df['kehujushou'] = df['kehujushou'].astype(float)
    df['wuhuojine'] = clean2f(df['wuhuojine'])
    df['wuhuojine'] = df['wuhuojine'].astype(float)
    df['shaopeijine'] = clean2f(df['shaopeijine'])
    df['shaopeijine'] = df['shaopeijine'].astype(float)
    df['peicuoweiyao'] = clean2f(df['peicuoweiyao'])
    df['peicuoweiyao'] = df['peicuoweiyao'].astype(float)
    # df['金额核对'] = clean(df['金额核对'])
    # df['金额核对'] = df['金额核对'].astype(float)
    # print(df.columns)
    # print(df[0:15])

    return df

# %% [markdown]
# df = pd.read_excel('2017年全单统计管理.xlsm',sheetname='全单统计管理',na_values=['NA'])
# descdb(df)

# %%
cnx = lite.connect('quandan.db')
# sql_df=df.loc[:,['订单日期', '配货人', '配货准确', '业务主管', '终端编码', '终端名称', '积欠', '送货金额',
#                  '实收金额', '收款方式', '退货金额', '客户拒收', '无货金额', '少配金额', '配错未要',
#                  '送达日期', '车辆', '送货人', '收款日期', '收款人', '拒收品项']]
# df.to_sql(name='sources', con=cnx, schema=sql_df, if_exists='replace', chunksize=10000)

# %%
def yingdacal(dfy,cnx):
    df = []
    dfjiaqi = pd.read_sql_query('select date from jiaqi',cnx)
    for x in dfy:
        if type(x) == str:
            x = pd.to_datetime(x)
        datestr = (x+pd.DateOffset(days=1)).strftime('%Y-%m-%d')

        # print(dfall.columns)
        # print(dfall['tianshu'])
        # print(len(dfall))
        # print(int(x.strftime('%w')))
        if(datestr in dfjiaqi['date']):
            dfjiaqitianshu = pd.read_sql_query('select tianshu from jiaqi where date =\'' + datestr + '\'', cnx)
            ii = x+pd.DateOffset(days=int(dfjiaqitianshu['tianshu'][0]))
        elif(int(x.strftime('%w')) == 6):
            ii = x+pd.DateOffset(days=2)
        else:
            ii = x + pd.DateOffset(days=1)
        df.append(ii)

    return df

# %%
df = pd.read_sql_query("select * from quandan",cnx)
# df = df[(df['peihuoren'] != '作废') &(df['dingdanriqi'] >= '2016-03-29')]
descdb(df)
df['应达日期'] = yingdacal(df['订单日期'])
descdb(df)

# %%
df['送达天数'] = (df['送达日期'] - df['订单日期']).dt.days
df['收款天数'] = (df['收款日期'] - df['送达日期']).dt.days
df = df[((df['送达天数'] >= -1)&(df['送达天数'] <= 14))]

# %%
descdb(df)
# df = df[['dingdanriqi','yewuzhuguan', 'zhongduanbianma', 'songhuojine', 'shishoujine', 'yingdariqi','songdariqi', 'songhuoren', 'shoukuanriqi', 'songdatianshu', 'shoukuantianshu']]
# # descdb(df)
# print(len(df))
# df = df.drop_duplicates()
# print(len(df))
# # df = df[(df['songdatianshu'] <= -1)]
# # # df = df[(df['shoukuantianshu'] <= -1 ) & (df['shoukuantianshu'] >-14000 )]
# # descdb(df)
# # print(df)
# # dfsong1 = df[(df['songdatianshu'] == 2)]
# # print(dfsong1)
# # dfsong2 = df[(df['songdatianshu'] > -1)]
# # print(dfsong2)
# # print(df['yewuzhuguan'].count())
# # dfsongcount = df.groupby('dingdanriqi')['songdatianshu'].count
# # print(dfsongcount)
#
#
# sql_df=df.loc[:,['dingdanriqi','yewuzhuguan', 'zhongduanbianma', 'songhuojine', 'shishoujine', 'songdariqi', 'songhuoren', 'shoukuanriqi', 'songdatianshu', 'shoukuantianshu']]
# df.to_sql(name='sources', con=cnx, schema=sql_df, if_exists='replace', chunksize=1000)
# dfall = pd.read_sql_query('select dingdanriqi, count(dingdanriqi) as zongdan from sources group by dingdanriqi', cnx)
# dfall.index = dfall['dingdanriqi']
# # descdb(dfall)
#
# dfnormal = pd.read_sql_query('select dingdanriqi, count(dingdanriqi) as zhengchangdan from sources where songdatianshu <= 3 group by dingdanriqi', cnx)
# dfnormal.index = dfnormal['dingdanriqi']
# descdb(dfnormal)
# dfabnormal = pd.read_sql_query('select dingdanriqi, count(dingdanriqi) as buzhengchangdan from sources where songdatianshu > 3 group by dingdanriqi', cnx)
# dfabnormal.index = dfabnormal['dingdanriqi']
# descdb(dfabnormal)
# cnx.close()
#
# plt.title(u'每日订单配送图')
# fig, ax1 = plt.subplots()
# # plt.figure(figsize=(16, 12))
# plt.plot(pd.to_datetime(dfall['dingdanriqi']), dfall['zongdan'], lw=1.5, label=u'总单')
# plt.plot(pd.to_datetime(dfnormal['dingdanriqi']), dfnormal['zhengchangdan'], lw=1.5, label=u'正常送达')
# # plt.plot(df['date'], df['fengsu'], lw=1.5, label=u'风速')
# plt.xlabel(u'日期')
# plt.ylabel(u'单数')
# # plt.axis('tight')
# plt.grid(True)
# plt.legend(loc=0)
# ax2 = ax1.twinx()
# plt.plot(pd.to_datetime(dfabnormal['dingdanriqi']), dfabnormal['buzhengchangdan'], 'c-', lw=1.5, label=u'未正常送达')
# plt.legend(loc=7)
# plt.ylabel(u'单数')
# plt.grid(True)
# plt.savefig('peisong.png')
# plt.close()
